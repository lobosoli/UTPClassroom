from __future__ import print_function

import pickle
import os.path
import json
import time
import sys
import openpyxl

from pprint import pprint
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request



# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/classroom.courses.readonly', 
        'https://www.googleapis.com/auth/classroom.rosters',
        'https://www.googleapis.com/auth/classroom.rosters.readonly',
        'https://www.googleapis.com/auth/classroom.profile.emails',
        'https://www.googleapis.com/auth/classroom.profile.photos',
	    'https://www.googleapis.com/auth/drive', 
        'https://www.googleapis.com/auth/spreadsheets',
	    'https://www.googleapis.com/auth/spreadsheets.readonly']

def check_auth(api, version):
    creds = None
    
    """Shows basic usage of the Classroom API.
    Prints the names of the first 10 courses the user has access to.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)

    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'estadisticasCR_id.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build(api, version, credentials=creds)   
    return service

def main():    
    
    # Call the Classroom API
    classroom_service = check_auth('classroom','v1') 
        
    results = classroom_service.courses().list(pageSize = 500).execute()
    courses = results.get('courses', [])
       
    # Number of rows   
    contar= 1

    if os.path.isfile("ClassRoomNumStudents.xlsx"):
        print("ERROR: El Archivo ya existe")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cursos"

        ws.cell(column = 1, row = 1, value ="{0}".format(str("ID Curso")))
        ws.cell(column = 2, row = 1, value ="{0}".format(str("Nombre Curso")))
        ws.cell(column = 3, row = 1, value ="{0}".format(str("ID Profesor")))
        ws.cell(column = 4, row = 1, value ="{0}".format(str("Nombre Profesor")))
        ws.cell(column = 5, row = 1, value ="{0}".format(str("Correo Profesor")))
        ws.cell(column = 6, row = 1, value ="{0}".format(str("Fecha de Creación")))
        ws.cell(column = 7, row = 1, value ="{0}".format(str("Última Actualización")))
        ws.cell(column = 8, row = 1, value ="{0}".format(str("Número de Estudiantes")))
        ws.cell(column = 9, row = 1, value ="{0}".format(str("Estado del Curso")))
    
    if not courses:
        print('No courses found.')
    else:               
        #print('ID Curso;Nombre Curso;ID Profesor;Nombre Profesor;Correo Profesor;Fecha de Creación;Ultima Actualización;Número de Estudiantes;Estado del Curso')
        page_token = results.get('nextPageToken',None)            
        while page_token != "":
            for course in courses:
                #idcurso = course['id']
                #idprofesor = course.get('ownerId')
                teacherID = classroom_service.userProfiles().get(userId = course.get('ownerId'), x__xgafv='2', alt="json").execute()
                Estudiantes = classroom_service.courses().students().list(courseId = course['id'], pageSize = 0, x__xgafv='2', alt="json").execute()
                datosEstudiantes = Estudiantes.get('students',[])
                estuCont = 0
                contar=contar + 1
                # Get the numbers of students
                if not datosEstudiantes:
                    estuCont = 0
                else:
                    for listaEstudiantes in datosEstudiantes:                            
                        estuCont = estuCont + 1
                #print(course['id'],';', course['name'],';', course['ownerId'],';', teacherID['name']['fullName'],';', teacherID['emailAddress'],';', course['creationTime'],';', course['updateTime'],';', estuCont ,';', course['courseState'])                              

                ws.cell(column = 1, row = contar, value ="{0}".format(str(course['id'])))
                ws.cell(column = 2, row = contar, value ="{0}".format(str(course['name'])))
                ws.cell(column = 3, row = contar, value ="{0}".format(str(course['ownerId'])))
                ws.cell(column = 4, row = contar, value ="{0}".format(str(teacherID['name']['fullName'])))
                ws.cell(column = 5, row = contar, value ="{0}".format(str(teacherID['emailAddress'])))
                ws.cell(column = 6, row = contar, value ="{0}".format(str(course['creationTime'])))
                ws.cell(column = 7, row = contar, value ="{0}".format(str(course['updateTime'])))
                ws.cell(column = 8, row = contar, value ="{0}".format(str(estuCont)))
                ws.cell(column = 9, row = contar, value ="{0}".format(str(course['courseState'])))
                wb.save(filename = "ClassRoomNumStudents.xlsx")                    
                
            page_token = results.get('nextPageToken',None)
            results = classroom_service.courses().list(pageToken = page_token, pageSize = 10).execute()
            courses = results.get('courses', [])
         

if __name__ == '__main__':
    main()
